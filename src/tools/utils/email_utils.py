import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import os
import tempfile
from datetime import datetime, date, time
from pathlib import Path
import jinja2
from dotenv import load_dotenv
import yfinance as yf 
from src.tools.utils.technical_analysis_utils import calculate_roc
from zoneinfo import ZoneInfo

TIMEZONE_STR = os.getenv("TIMEZONE", "Asia/Kolkata")
IST = ZoneInfo(TIMEZONE_STR)
# Load environment variables
load_dotenv()

def make_html_template(data_list: list, title: str = "Stock Predictions"):
    """
    Generates an HTML report from a list of stock data dictionaries using Jinja2.
    """
    print(f"📄 [Email Utils] Generating HTML template for {len(data_list)} items.")
    try:
        template_str = """
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body { font-family: Arial, sans-serif; }
                table { width: 100%; border-collapse: collapse; margin-top: 20px; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; }
                h1 { color: #333; }
                .bullish { color: green; font-weight: bold; }
                .bearish { color: red; font-weight: bold; }
            </style>
        </head>
        <body>
            <h1>{{ title }}</h1>
            <p>Total Stocks Found: <strong>{{ data|length }}</strong></p>
            <table>
                <thead>
                    <tr>
                        <th>Ticker</th>
                        <th>Price</th>
                        <th>Trend</th>
                        <th>Patterns</th>
                        <th>RSI</th>
                        <th>MACD</th>
                    </tr>
                </thead>
                <tbody>
                    {% for item in data %}
                    <tr>
                        <td>{{ item.ticker }}</td>
                        <td>{{ item.price }}</td>
                        <td class="{{ 'bullish' if item.trend == 1 else 'bearish' }}">
                            {{ 'Bullish' if item.trend == 1 else 'Bearish' }}
                        </td>
                        <td>{{ item.patterns if item.patterns else '-' }}</td>
                        <td>{{ "%.2f"|format(item.indicators.rsi) if item.indicators.rsi else 'N/A' }}</td>
                        <td>{{ "%.2f"|format(item.indicators.macd) if item.indicators.macd else 'N/A' }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </body>
        </html>
        """
        template = jinja2.Template(template_str)
        html_output = template.render(data=data_list, title=title)
        print("✅ [Email Utils] HTML generated successfully.")
        return html_output
    except Exception as e:
        print(f"❌ [Email Utils] Error generating HTML: {e}")
        return "<html><body><h1>Error generating report</h1></body></html>"

def send_mail(recipient: str, subject: str, body: str, attachment_path: str = None):
    """
    Sends an email with an optional attachment.
    Fetches credentials from environment variables.
    """
    print(f"📧 [Email Utils] Preparing to send email to {recipient}...")
    
    # Get config from env
    sender_email = os.getenv("EMAIL_USER")
    sender_password = os.getenv("EMAIL_PASSWORD")
    smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port_str = os.getenv("SMTP_PORT", "587")
    
    if not sender_email or not sender_password:
        print("⚠️ [Email Utils] EMAIL_USER or EMAIL_PASSWORD not set in .env. Mocking email send.")
        print(f"   [Mock] To: {recipient}, Subject: {subject}, Attachment: {attachment_path}")
        return True

    try:
        smtp_port = int(smtp_port_str)
        
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient
        msg['Subject'] = subject

        msg.attach(MIMEText(body, 'html'))

        if attachment_path and os.path.exists(attachment_path):
            print(f"📎 [Email Utils] Attaching file: {attachment_path}")
            with open(attachment_path, "rb") as f:
                part = MIMEApplication(f.read(), Name=os.path.basename(attachment_path))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(attachment_path)}"'
            msg.attach(part)
        elif attachment_path:
             print(f"⚠️ [Email Utils] Attachment file not found: {attachment_path}")

        print(f"🔌 [Email Utils] Connecting to SMTP server {smtp_server}:{smtp_port}...")
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()
        print(f"✅ [Email Utils] Email sent successfully to {recipient}")
        return True
    
    except Exception as e:
        print(f"❌ [Email Utils] Failed to send email: {e}")
        return False


# =============================================================================
# Workflow-specific helpers (used by workflow_executor.py)
# =============================================================================

def make_workflow_html(stocks: list, title: str = "Strend Workflow Report") -> str:
    """
    Generate a rich HTML report for the workflow pipeline results.
    Includes stock data, technical indicators, fundamental scores, and news.
    """
    print(f"📄 [Email Utils] Generating workflow HTML for {len(stocks)} stocks.")
    try:
        template_str = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            font-family: 'Segoe UI', Arial, sans-serif;
            background: #f4f6f9;
            color: #333;
            padding: 30px;
        }
        .container { max-width: 1000px; margin: 0 auto; }
        .header {
            background: linear-gradient(135deg, #1a1a2e, #16213e, #0f3460);
            color: #fff;
            padding: 30px 40px;
            border-radius: 12px;
            margin-bottom: 24px;
        }
        .header h1 { font-size: 28px; margin-bottom: 8px; }
        .header p { font-size: 14px; opacity: 0.85; }
        .summary-bar {
            display: flex;
            gap: 16px;
            margin-bottom: 24px;
            flex-wrap: wrap;
        }
        .summary-card {
            background: #fff;
            border-radius: 10px;
            padding: 16px 24px;
            flex: 1;
            min-width: 140px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            text-align: center;
        }
        .summary-card .label { font-size: 12px; color: #888; text-transform: uppercase; }
        .summary-card .value { font-size: 24px; font-weight: 700; color: #1a1a2e; margin-top: 4px; }
        table {
            width: 100%;
            border-collapse: collapse;
            background: #fff;
            border-radius: 10px;
            overflow: hidden;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            margin-bottom: 24px;
        }
        th {
            background: #1a1a2e;
            color: #fff;
            padding: 12px 14px;
            text-align: left;
            font-size: 12px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        td {
            padding: 10px 14px;
            border-bottom: 1px solid #eee;
            font-size: 13px;
        }
        tr:last-child td { border-bottom: none; }
        tr:hover { background: #f8f9fc; }
        .bullish { color: #16a34a; font-weight: 600; }
        .bearish { color: #dc2626; font-weight: 600; }
        .neutral { color: #888; }
        .badge {
            display: inline-block;
            padding: 3px 10px;
            border-radius: 20px;
            font-size: 11px;
            font-weight: 600;
        }
        .badge-green { background: #dcfce7; color: #16a34a; }
        .badge-red { background: #fee2e2; color: #dc2626; }
        .badge-yellow { background: #fef9c3; color: #a16207; }
        .badge-blue { background: #dbeafe; color: #2563eb; }
        .news-block { margin-top: 6px; }
        .news-item {
            font-size: 12px;
            color: #555;
            padding: 3px 0;
            border-left: 3px solid #3b82f6;
            padding-left: 8px;
            margin-bottom: 4px;
        }
        .news-time { color: #999; font-size: 11px; }
        .footer {
            text-align: center;
            font-size: 12px;
            color: #aaa;
            padding: 20px;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{{ title }}</h1>
            <p>Generated on {{ today }} | {{ count }} stock(s) in final results</p>
        </div>

        <div class="summary-bar">
            <div class="summary-card">
                <div class="label">Total Stocks</div>
                <div class="value">{{ count }}</div>
            </div>
            {% if has_technical %}
            <div class="summary-card">
                <div class="label">Bullish</div>
                <div class="value" style="color:#16a34a">{{ bullish_count }}</div>
            </div>
            <div class="summary-card">
                <div class="label">Bearish</div>
                <div class="value" style="color:#dc2626">{{ bearish_count }}</div>
            </div>
            {% endif %}
            {% if has_news %}
            <div class="summary-card">
                <div class="label">With News</div>
                <div class="value" style="color:#2563eb">{{ news_count }}</div>
            </div>
            {% endif %}
        </div>

        <table>
            <thead>
                <tr>
                    <th>#</th>
                    <th>Ticker</th>
                    <th>Name</th>
                    <th>Price</th>
                    <th>Volume</th>
                    <th>% Chg</th>
                    {% if has_technical %}
                    <th>Trend</th>
                    <th>RSI</th>
                    <th>MACD</th>
                    <th>ADX</th>
                    <th>ROC (5m)</th>
                    {% endif %}
                    {% if has_fundamental %}
                    <th>Fund. Score</th>
                    <th>Rating</th>
                    {% endif %}
                </tr>
            </thead>
            <tbody>
                {% for stock in stocks %}
                <tr>
                    <td>{{ loop.index }}</td>
                    <td><strong>{{ stock.ticker }}</strong></td>
                    <td>{{ stock.name }}</td>
                    <td>₹{{ "%.2f"|format(stock.price) if stock.price else 'N/A' }}</td>
                    <td>{{ "{:,.0f}".format(stock.volume) if stock.volume else 'N/A' }}</td>
                    <td class="{{ 'bullish' if stock.percent_change > 0 else 'bearish' if stock.percent_change < 0 else 'neutral' }}">
                        {{ "%.2f"|format(stock.percent_change) if stock.percent_change else '0.00' }}%
                    </td>
                    {% if has_technical %}
                    <td>
                        {% if stock.trend == 1 %}
                            <span class="badge badge-green">Bullish</span>
                        {% else %}
                            <span class="badge badge-red">Bearish</span>
                        {% endif %}
                    </td>
                    <td>{{ "%.1f"|format(stock.indicators.rsi) if stock.indicators else 'N/A' }}</td>
                    <td>{{ "%.2f"|format(stock.indicators.macd) if stock.indicators else 'N/A' }}</td>
                    <td>{{ "%.1f"|format(stock.indicators.adx) if stock.indicators else 'N/A' }}</td>
                    <td>{{ "%.2f"|format(stock.indicators.roc) if stock.indicators and stock.indicators.roc else '0.00' }}</td>
                    {% endif %}
                    {% if has_fundamental %}
                    <td>
                        {% if stock.fundamental_data %}
                            <span class="badge badge-blue">{{ "%.1f"|format(stock.fundamental_data.score_percentage) }}%</span>
                        {% else %}
                            N/A
                        {% endif %}
                    </td>
                    <td>{{ stock.fundamental_data.rating if stock.fundamental_data else 'N/A' }}</td>
                    {% endif %}
                </tr>
                {% if has_news and stock.news %}
                <tr>
                    <td></td>
                    <td colspan="{{ total_cols - 1 }}">
                        <div class="news-block">
                        {% for n in stock.news[:3] %}
                            <div class="news-item">
                                {{ n.news }} <span class="news-time">— {{ n.time_str }}</span>
                            </div>
                        {% endfor %}
                        </div>
                    </td>
                </tr>
                {% endif %}
                {% endfor %}
            </tbody>
        </table>

        <div class="footer">
            Strend Workflow Report — Generated automatically
        </div>
    </div>
</body>
</html>
        """
        # Compute template vars
        has_technical = any(s.get("indicators") for s in stocks)
        has_fundamental = any(s.get("fundamental_data") for s in stocks)
        has_news = any(s.get("news") for s in stocks)

        bullish_count = sum(1 for s in stocks if s.get("trend") == 1)
        bearish_count = sum(1 for s in stocks if s.get("trend") == 0 and "trend" in s)
        news_count = sum(1 for s in stocks if s.get("news"))

        total_cols = 6  # base columns
        if has_technical:
            total_cols += 5
        if has_fundamental:
            total_cols += 2

        # Make stocks accessible as objects for Jinja2 dot notation
        class DotDict(dict):
            __getattr__ = dict.get
            __setattr__ = dict.__setitem__

        safe_stocks = []
        for s in stocks:
            ds = DotDict(s)
            if "indicators" in s and isinstance(s["indicators"], dict):
                ds["indicators"] = DotDict(s["indicators"])
            if "fundamental_data" in s and isinstance(s["fundamental_data"], dict):
                ds["fundamental_data"] = DotDict(s["fundamental_data"])
            safe_stocks.append(ds)

        template = jinja2.Template(template_str)
        html_output = template.render(
            title=title,
            today=date.today().isoformat(),
            count=len(stocks),
            stocks=safe_stocks,
            has_technical=has_technical,
            has_fundamental=has_fundamental,
            has_news=has_news,
            bullish_count=bullish_count,
            bearish_count=bearish_count,
            news_count=news_count,
            total_cols=total_cols,
        )
        print("✅ [Email Utils] Workflow HTML generated successfully.")
        return html_output

    except Exception as e:
        print(f"❌ [Email Utils] Error generating workflow HTML: {e}")
        return f"<html><body><h1>Error generating report</h1><p>{e}</p></body></html>"


def html_to_pdf(html_content: str, title: str = "report") -> str:
    """
    Convert HTML string to a PDF file using WeasyPrint.
    Returns the path to the generated temporary PDF file.
    """
    print("📄 [Email Utils] Converting HTML to PDF with WeasyPrint...")
    try:
        from weasyprint import HTML

        safe_title = "".join(c if c.isalnum() or c in "-_ " else "" for c in title).strip()
        filename = f"{safe_title}_{date.today().isoformat()}.pdf"
        filepath = os.path.join(tempfile.gettempdir(), filename)

        HTML(string=html_content).write_pdf(filepath)
        print(f"✅ [Email Utils] PDF generated: {filepath}")
        return filepath
    except ImportError:
        print("⚠️ [Email Utils] WeasyPrint not installed. Falling back to HTML attachment.")
        # Fallback: save as .html
        filename = f"{title}_{date.today().isoformat()}.html"
        filepath = os.path.join(tempfile.gettempdir(), filename)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html_content)
        return filepath
    except Exception as e:
        print(f"❌ [Email Utils] PDF generation failed: {e}")
        return None


def stocks_to_excel(stocks: list, title: str = "report") -> str:
    """
    Convert a list of stock dicts to an Excel file using openpyxl.
    Returns the path to the generated temporary Excel file.
    """
    print(f"📊 [Email Utils] Generating Excel report for {len(stocks)} stocks...")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        safe_title = "".join(c if c.isalnum() or c in "-_ " else "" for c in title).strip()
        filename = f"{safe_title}_{date.today().isoformat()}.xlsx"
        filepath = os.path.join(tempfile.gettempdir(), filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workflow Results"

        # Determine which columns to include
        has_technical = any(s.get("indicators") for s in stocks)
        has_fundamental = any(s.get("fundamental_data") for s in stocks)
        has_news = any(s.get("news") for s in stocks)

        # Build headers
        col_headers = ["#", "Ticker", "Name", "Price", "TodayHigh", "TodayLow", "52weekHigh", "Volume", "% Change", "ROC (5-5)","ROC Today", "Diff ROC"]
        if has_technical:
            col_headers.extend(["Trend", "RSI", "MACD", "MACD Signal", "MACD Hist", "ADX"])
        if has_fundamental:
            col_headers.extend(["Fund. Score %", "Rating", "Risk Level"])
        if has_news:
            col_headers.extend(["News Headlines", "News Time"])

        ws.append(col_headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for idx, stock in enumerate(stocks, 1):
            indicators = stock.get("indicators", {})
            fund = stock.get("fundamental_data", {})
            news_list = stock.get("news", [])
            s = yf.Ticker(f"{stock['ticker']}.NS")
            week_high = s.info.get("fiftyTwoWeekHigh", None)
            hist = s.history(period="1d")
            if not hist.empty:
                today_high = hist['High'].iloc[0]
                today_low = hist['Low'].iloc[0]
            else:
                today_high = None
                today_low = None
            
            roc = calculate_roc(stock["ticker"], 12, "5m")
            
            today = datetime.now(IST).date()

            #Create a timezone-aware datetime for 9:15 AM on today
            start_time = datetime.combine(today, time(9, 15), tzinfo=IST)

            #Get current time as timezone-aware datetime
            current_time = datetime.now(IST)

            # Calculate period in minutes (300 seconds = 5 minutes)
            minutes_diff = (current_time - start_time).total_seconds() // 300
            period = max(1, int(minutes_diff))
            today_roc_value = calculate_roc(stock["ticker"], period=period, interval="5m")

            diff_roc = today_roc_value - roc if today_roc_value is not None and roc is not None else 0

            row = [
                idx,
                stock.get("ticker", ""),
                stock.get("name", ""),
                stock.get("price", 0),
                today_high,
                today_low,
                week_high,
                stock.get("volume", 0),
                stock.get("percent_change", 0),
                round(roc, 2) if roc is not None else 0,
                round(today_roc_value, 2) if today_roc_value is not None else 0,
                round(diff_roc, 2),
            ]

            if has_technical:
                trend_str = "Bullish" if stock.get("trend") == 1 else "Bearish"
                row.extend([
                    trend_str,
                    round(indicators.get("rsi", 0), 2),
                    round(indicators.get("macd", 0), 2),
                    round(indicators.get("macd_signal", 0), 2),
                    round(indicators.get("macd_hist", 0), 2),
                    round(indicators.get("adx", 0), 2),
                    round(indicators.get("roc", 0), 2),
                ])

            if has_fundamental:
                row.extend([
                    round(fund.get("score_percentage", 0), 1),
                    fund.get("rating", "N/A"),
                    fund.get("risk_level", "N/A"),
                ])

            if has_news:
                headlines = "\n".join(n.get("news", "")[:100] for n in news_list[:5])
                times = "\n".join(n.get("time_str", "") for n in news_list[:5])
                row.extend([headlines, times])

            ws.append(row)

        # Color trend column
        if has_technical:
            trend_col_idx = 7  # 1-indexed
            green_font = Font(color="16A34A", bold=True)
            red_font = Font(color="DC2626", bold=True)
            for row_num in range(2, len(stocks) + 2):
                cell = ws.cell(row=row_num, column=trend_col_idx)
                if cell.value == "Bullish":
                    cell.font = green_font
                else:
                    cell.font = red_font

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = max(len(line) for line in str(cell.value or "").split("\n"))
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        wb.save(filepath)
        print(f"✅ [Email Utils] Excel generated: {filepath}")
        return filepath

    except Exception as e:
        print(f"❌ [Email Utils] Excel generation failed: {e}")
        return None
