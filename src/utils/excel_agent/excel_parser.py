"""
Excel Parser
=============
Single Responsibility: Parse Excel data (JSON or bytes) into structured rows.
Does NOT touch databases or LLMs.
"""

import io
import json
import re
from typing import Tuple, List, Dict, Any


def normalize_column_name(name: str) -> str:
    """
    Normalize a column name to be SQLite-safe.
    
    Examples:
        'Stock Price (%)'  → 'stock_price_pct'
        'ROC  Change'      → 'roc_change'
        '52W High'         → '_52w_high'
        'Column Name!'     → 'column_name'
    """
    if not name or not isinstance(name, str):
        return "unnamed"

    s = name.strip().lower()
    s = s.replace("%", "pct").replace("&", "and")
    s = re.sub(r"[^a-z0-9_]", "_", s)    # Replace non-alphanumeric with _
    s = re.sub(r"_+", "_", s)              # Collapse multiple underscores
    s = s.strip("_")

    # SQLite columns can't start with a digit
    if s and s[0].isdigit():
        s = "_" + s

    return s or "unnamed"


def normalize_columns(columns: List[str]) -> List[str]:
    """
    Normalize all column names, handling duplicates by appending suffix.
    
    Args:
        columns: Raw column names from Excel
        
    Returns:
        List of unique, normalized column names
    """
    normalized = []
    seen = {}

    for col in columns:
        base = normalize_column_name(col)
        if base in seen:
            seen[base] += 1
            base = f"{base}_{seen[base]}"
        else:
            seen[base] = 0
        normalized.append(base)

    return normalized


def parse_excel_json(data: dict) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse the {"file": "..."} JSON payload from WebSocket.
    
    Supports two JSON formats:
    1. Array of objects: [{"col1": "v1", "col2": "v2"}, ...]
    2. Array of arrays with first row as headers: [["col1","col2"], ["v1","v2"], ...]
    
    Args:
        data: Dict with "file" key containing JSON-serialized sheet data
        
    Returns:
        (columns, rows) where rows is a list of dicts with normalized column keys
        
    Raises:
        ValueError: If data format is invalid
    """
    file_content = data.get("file")
    if not file_content:
        raise ValueError("Missing 'file' key in payload")

    # Parse the JSON string if it's a string
    if isinstance(file_content, str):
        try:
            parsed = json.loads(file_content)
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON in 'file' field: {e}")
    elif isinstance(file_content, (list, dict)):
        parsed = file_content
    else:
        raise ValueError(f"Unexpected type for 'file': {type(file_content)}")

    if not isinstance(parsed, list) or len(parsed) == 0:
        raise ValueError("'file' must be a non-empty JSON array")

    # Format 1: Array of objects [{col: val}, ...]
    if isinstance(parsed[0], dict):
        raw_columns = list(parsed[0].keys())
        columns = normalize_columns(raw_columns)
        col_map = dict(zip(raw_columns, columns))

        rows = []
        for row_data in parsed:
            rows.append({col_map[k]: v for k, v in row_data.items() if k in col_map})
        return columns, rows

    # Format 2: Array of arrays [["header1","header2"], ["val1","val2"], ...]
    if isinstance(parsed[0], list):
        if len(parsed) < 2:
            raise ValueError("Array-of-arrays format needs at least a header row + 1 data row")

        raw_columns = [str(c) for c in parsed[0]]
        columns = normalize_columns(raw_columns)

        rows = []
        for row_arr in parsed[1:]:
            row_dict = {}
            for i, val in enumerate(row_arr):
                if i < len(columns):
                    row_dict[columns[i]] = val
            rows.append(row_dict)
        return columns, rows

    raise ValueError(f"Unsupported row format: {type(parsed[0])}")


def parse_excel_bytes(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse raw .xlsx bytes via openpyxl.
    
    Args:
        file_bytes: Raw bytes of an .xlsx file
        
    Returns:
        (columns, rows) with normalized column names
        
    Raises:
        ValueError: If file cannot be parsed
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            raise ValueError("Excel file needs at least a header row + 1 data row")

        raw_columns = [str(c) if c else f"col_{i}" for i, c in enumerate(all_rows[0])]
        columns = normalize_columns(raw_columns)

        rows = []
        for row_tuple in all_rows[1:]:
            row_dict = {}
            for i, val in enumerate(row_tuple):
                if i < len(columns):
                    row_dict[columns[i]] = val
            rows.append(row_dict)

        wb.close()
        return columns, rows

    except Exception as e:
        raise ValueError(f"Failed to parse Excel file: {e}")
