"""
Strategy Scheduler — Runs at configurable times (Mon-Fri) to:
1. Scrape stock data from Chartink screener
2. Apply price/volume filters (volume threshold increases with each time slot)
3. Scrape news from Groww for filtered stocks
4. Keep only stocks with recent/today's news
5. Send notifications to frontend via WebSocket
6. Cache results in Excel files (auto-cleaned daily)

All configuration is loaded from .env with sensible defaults.
"""

import os
import sys
import asyncio
import logging
import re
from datetime import datetime, date, time
from pathlib import Path
from dotenv import load_dotenv
import yfinance as yf
from src.tools.utils.email_utils import send_mail, make_workflow_html
from src.core.db import get_db_connection
from zoneinfo import ZoneInfo

# Ensure project root is in sys.path for absolute imports
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
if project_root not in sys.path:
    sys.path.append(project_root)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("strategy_scheduler")

# Load .env from project root
load_dotenv(os.path.join(project_root, ".env"))

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

from src.tools.utils.chartink_scraper import fetch_chartink_data
from src.tools.utils.news_scraper import scrape_news_from_groww
from src.core.notification_manager import notification_manager
from src.tools.utils.technical_analysis_utils import calculate_roc


# =============================================================================
# Configuration (loaded from .env with defaults)
# =============================================================================

CHARTINK_URL = os.getenv(
    "STRATEGY_CHARTINK_URL",
    "https://chartink.com/screener/macd-bearish-or-bullish-crossover",
)
CHARTINK_QUERY = os.getenv(
    "STRATEGY_CHARTINK_QUERY",
    "RSI(14) > 60 AND Volume > 1.5× 20-day avg AND price above 200 EMA AND price below 52W high by less than 5%.",
)
CHARTINK_PAGES = int(os.getenv("STRATEGY_CHARTINK_PAGES", "5"))

# Filter thresholds
MIN_PRICE = float(os.getenv("STRATEGY_MIN_PRICE", "100"))

# Volume thresholds per schedule slot — increases as the trading day progresses.
# Comma-separated list matching each schedule slot in order.
# Default: 50000 for 9:34, 100000 for 9:51, 200000 for 10:05, 500000 for 10:30
VOLUME_THRESHOLDS_STR = os.getenv("STRATEGY_VOLUME_THRESHOLDS", "50000,100000,200000,500000")

# Schedule times (HH:MM, comma-separated) — Mon–Fri only
SCHEDULE_TIMES_STR = os.getenv("STRATEGY_SCHEDULE_TIMES", "9:34,9:51,10:05,10:30")

# Timezone Configuration
TIMEZONE_STR = os.getenv("TIMEZONE", "Asia/Kolkata")
IST = ZoneInfo(TIMEZONE_STR)

# Cache directory for daily Excel exports
CACHE_DIR = os.path.join(project_root, "cache", "strategy")

# Patterns that indicate "recent" / "today's" news
RECENT_NEWS_PATTERNS = [
    r"\d+\s*sec(?:s|ond|onds)?\s*ago",
    r"\d+\s*min(?:s|ute|utes)?\s*ago",
    r"\d+\s*hour(?:s)?\s*ago",
    r"a\s+min(?:ute)?\s+ago",
    r"a\s+sec(?:ond)?\s+ago",
    r"an?\s+hour\s+ago",
    r"just\s*now",
    r"a\s+day\s+ago",
    r"1\s+day\s+ago",
    r"today",
]

RECENT_NEWS = os.getenv("RECENT_NEWS", "True")

# =============================================================================
# Helper Functions
# =============================================================================


def _parse_volume_thresholds() -> list[float]:
    """Parse comma-separated volume thresholds from env string."""
    return [float(v.strip()) for v in VOLUME_THRESHOLDS_STR.split(",") if v.strip()]


def get_volume_threshold_for_slot(slot_index: int) -> float:
    """Return minimum volume for a schedule slot. Falls back to last threshold."""
    thresholds = _parse_volume_thresholds()
    if slot_index < len(thresholds):
        return thresholds[slot_index]
    return thresholds[-1] if thresholds else 50000


def is_recent_news(time_str: str) -> bool:
    """Check if a time string indicates recent / today's news."""
    time_lower = time_str.lower().strip()
    return any(re.search(p, time_lower) for p in RECENT_NEWS_PATTERNS)


def parse_number(value: str) -> float:
    """Safely parse a formatted number string (handles commas, spaces, etc.)."""
    try:
        cleaned = re.sub(r"[^\d.\-]", "", str(value))
        return float(cleaned) if cleaned else 0.0
    except (ValueError, TypeError):
        return 0.0


def find_column_index(headers: list[str], *keywords: str) -> int:
    """Find the first column index whose header matches any of the keywords."""
    for i, header in enumerate(headers):
        h = header.lower().strip()
        for kw in keywords:
            if kw.lower() in h:
                return i
    return -1


# =============================================================================
# Cache Management
# =============================================================================


def clean_old_cache():
    """Remove Excel cache files from previous days."""
    today_str = datetime.now(IST).date().isoformat()
    cache_path = Path(CACHE_DIR)

    if not cache_path.exists():
        return

    for f in cache_path.iterdir():
        if f.is_file() and f.suffix == ".xlsx" and today_str not in f.name:
            logger.info(f"🗑️  Removing old cache: {f.name}")
            try:
                f.unlink()
            except OSError as e:
                logger.warning(f"⚠️  Could not delete {f.name}: {e}")


def save_to_excel(data: list[dict], slot_label: str) -> str:
    """Save filtered stock data with news to an Excel file. Returns filepath."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    cache_path = Path(CACHE_DIR)
    cache_path.mkdir(parents=True, exist_ok=True)

    today_str = datetime.now(IST).date().isoformat()
    timestamp = datetime.now(IST).strftime("%H%M")
    filename = f"strategy_{today_str}_{slot_label}_{timestamp}.xlsx"
    filepath = cache_path / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filtered Stocks"

    # --- Headers ---
    col_headers = [
        "Stock Name",
        "Ticker",
        "Price",
        "todayHigh",
        "52WeekHigh",
        "Volume",
        "ROC_Today", 
        "ROC (5m, 12p)",
        "News Headlines",
        "News Time",
        "News URL",
    ]
    ws.append(col_headers)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # --- Data rows ---
    for stock in data:
        news_list = stock.get("news", [])
        news_headlines = "\n".join(n.get("news", "") for n in news_list)
        news_times = "\n".join(n.get("time_str", "") for n in news_list)
        news_urls = "\n".join(n.get("url", "") for n in news_list)

        ws.append([
            stock.get("name", ""),
            stock.get("ticker", ""),
            stock.get("price", 0),
            stock.get("todayHigh", 0),
            stock.get("52weekHigh", 0),
            stock.get("volume", 0),
            round(stock.get("today_roc", 0), 2),
            round(stock.get("roc", 0), 2),
            news_headlines,
            news_times,
            news_urls,
        ])

    # --- Auto-fit column widths ---
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = max(len(line) for line in str(cell.value or "").split("\n"))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb.save(filepath)
    logger.info(f"💾 Data saved to: {filepath}")
    return str(filepath)


# =============================================================================
# Main Strategy Job
# =============================================================================


async def run_strategy_job(slot_index: int, slot_label: str):

    EMAIL_RECIPIENT = []
    conn = get_db_connection()
    with conn.cursor() as cur:
        cur.execute('SELECT email_id FROM "user"')
        for row in cur:
            EMAIL_RECIPIENT.append(row[0])
    conn.close()
    """
    Main scheduled job:
    1. Fetch Chartink data
    2. Filter by price & volume
    3. Scrape news for surviving stocks
    4. Keep only those with recent news
    5. Save to Excel & broadcast notifications
    """
    logger.info(f"{'='*60}")
    logger.info(f"🚀 Strategy Job START — Slot: {slot_label} (index {slot_index})")
    logger.info(f"{'='*60}")

    # Step 0: Housekeeping — remove yesterday's cache
    clean_old_cache()

    min_volume = get_volume_threshold_for_slot(slot_index)
    logger.info(f"📊 Filters → Min Price: ₹{MIN_PRICE}, Min Volume: {min_volume:,.0f}")

    # -----------------------------------------------------------------
    # Step 1: Fetch data from Chartink
    # -----------------------------------------------------------------
    try:
        headers, all_rows = await fetch_chartink_data(
            url=CHARTINK_URL,
            query_text=CHARTINK_QUERY,
            total_pages=CHARTINK_PAGES,
        )
    except Exception as e:
        logger.error(f"❌ Chartink fetch failed: {e}", exc_info=True)
        return

    if not headers or not all_rows:
        logger.warning("⚠️  No data returned from Chartink. Exiting job.")
        return

    logger.info(f"📄 Chartink returned {len(all_rows)} rows | Headers: {headers}")
    print(f"\n--- ALL EXTRACTED STOCKS DATA ---\n{all_rows}\n{'='*30}\n")

    # -----------------------------------------------------------------
    # Step 2: Identify column indices
    # -----------------------------------------------------------------
    name_idx = find_column_index(headers, "stock", "name", "company")
    ticker_idx = find_column_index(headers, "symbol", "ticker", "nsecode")
    price_idx = find_column_index(headers, "close", "price", "ltp")
    volume_idx = find_column_index(headers, "volume", "vol")

    # Positional fallbacks if header matching fails
    if name_idx == -1 and len(headers) > 1:
        name_idx = 1
    if ticker_idx == -1 and len(headers) > 2:
        ticker_idx = 2
    if price_idx == -1 and len(headers) > 3:
        price_idx = 3
    if volume_idx == -1 and len(headers) > 4:
        volume_idx = 4

    logger.info(
        f"📌 Column mapping → Name:{name_idx}, Ticker:{ticker_idx}, "
        f"Price:{price_idx}, Volume:{volume_idx}"
    )

    # -----------------------------------------------------------------
    # Step 3: Apply price & volume filters
    # -----------------------------------------------------------------
    filtered_stocks: list[dict] = []

    for row in all_rows:
        try:
            name = row[name_idx].strip() if name_idx < len(row) else "Unknown"
            ticker = row[ticker_idx].strip() if ticker_idx < len(row) else ""
            price = parse_number(row[price_idx]) if price_idx < len(row) else 0
            volume = parse_number(row[volume_idx]) if volume_idx < len(row) else 0

            if price >= MIN_PRICE and volume >= min_volume:
                filtered_stocks.append({
                    "name": name,
                    "ticker": ticker,
                    "price": price,
                    "volume": volume,
                    "raw_row": row,
                })
        except (IndexError, ValueError) as e:
            logger.debug(f"⚠️  Row parse error: {e}")
            continue

    logger.info(
        f"🔍 Filter result: {len(filtered_stocks)}/{len(all_rows)} stocks passed "
        f"(price≥{MIN_PRICE}, vol≥{min_volume:,.0f})"
    )
    print(f"\n--- FILTERED STOCKS DATA ---\n{filtered_stocks}\n{'='*30}\n")

    if not filtered_stocks:
        logger.info("📭 No stocks passed filters. Job finished.")
        return
    
    # -----------------------------------------------------------------
    # Step 4: Scrape news & keep only stocks with recent news
    # -----------------------------------------------------------------
    stocks_with_recent_news: list[dict] = []
    
    if RECENT_NEWS == 'True':
        for stock in filtered_stocks:
            ticker = stock["ticker"]
            if not ticker:
                continue

            logger.info(f"📰 Scraping news for: {ticker}")
            try:
                news_items = await scrape_news_from_groww(ticker)
            except Exception as e:
                logger.warning(f"⚠️  News scraping failed for {ticker}: {e}")
                continue

            # Keep only items whose time_str indicates recent / today
            recent_news = [n for n in news_items if is_recent_news(n.get("time_str", ""))]

            if recent_news:
                stock["news"] = recent_news
                stocks_with_recent_news.append(stock)
                logger.info(f"  ✅ {ticker} → {len(recent_news)} recent news item(s)")
            else:
                logger.info(f"  ⏭️  {ticker} → no recent news, skipped")

        logger.info(f"📬 Stocks with recent news: {len(stocks_with_recent_news)}")

        if not stocks_with_recent_news:
            logger.info("📭 No stocks with recent news. Job finished.")
            return
    else:
        stocks_with_recent_news = filtered_stocks

    #store 52week high value also
    for stock in stocks_with_recent_news:
        s = yf.Ticker(f"{stock["ticker"]}.NS")
        fifty_two_week_high = s.info.get('fiftyTwoWeekHigh')
        stock["52weekHigh"] = max(fifty_two_week_high, stock["price"])

        #high
        hist = s.history(period="1d")
    
        if not hist.empty:
            today_high = hist['High'].iloc[0]
        else:
            today_high = None
        
        stock["todayHigh"] = today_high

        # Calculate ROC (5m, 12 periods)
        try:
            roc_value = await asyncio.to_thread(calculate_roc, stock["ticker"], period=12, interval="5m")
            stock["roc"] = roc_value
            #start from 9:15 priod change according to time, like at 9:20 period = 2, 9:25 period = 3 etc

            #ERROR:2026-05-04 11:56:21,777 - strategy_scheduler - WARNING - ⚠️ ROC calculation failed for IDEAFORGE: can't subtract offsett-naive and offset-aware datetimes

            #Correct way to calculate roc for current time
            #Get today's date without time information
            today = datetime.now(IST).date()

            #Create a timezone-aware datetime for 9:15 AM on today
            start_time = datetime.combine(today, time(9, 15), tzinfo=IST)

            #Get current time as timezone-aware datetime
            current_time = datetime.now(IST)

            # Calculate period in minutes (300 seconds = 5 minutes)
            minutes_diff = (current_time - start_time).total_seconds() // 300
            period = max(1, int(minutes_diff))
            today_roc_value = await asyncio.to_thread(calculate_roc, stock["ticker"], period=period, interval="5m")
            stock["today_roc"] = today_roc_value
        except Exception as e:
            logger.warning(f"⚠️ ROC calculation failed for {stock['ticker']}: {e}")
            stock["roc"] = 0.0
    
    # -----------------------------------------------------------------
    # Step 5: Save results to Excel (daily cache)
    # -----------------------------------------------------------------
    try:
        excel_path = save_to_excel(stocks_with_recent_news, slot_label)
        logger.info(f"📁 Excel cached at: {excel_path}")
    except Exception as e:
        logger.error(f"❌ Excel save failed: {e}", exc_info=True)
    
    # -----------------------------------------------------------------
    # Step 6: Send email report
    # -----------------------------------------------------------------

    
    try:
        send_mail(
            recipient=",".join(EMAIL_RECIPIENT),
            subject=f"Strend Workflow Report - {slot_label} | {datetime.now(IST).date()}",
            body="Stocks from {slot_label} strategy",
            attachment_path=excel_path,
        )
    except Exception as e:
        logger.error(f"❌ Email sending failed: {e}", exc_info=True)
    
    # -----------------------------------------------------------------
    # Step 7: Broadcast notifications to frontend
    # -----------------------------------------------------------------
    for stock in stocks_with_recent_news:
        notification = {
            "type": "strategy_alert",
            "stock_name": stock["name"],
            "ticker": stock["ticker"],
            "price": stock["price"],
            "volume": stock["volume"],
            "news": [
                {
                    "headline": n.get("news", ""),
                    "time": n.get("time_str", ""),
                    "url": n.get("url", ""),
                }
                for n in stock.get("news", [])
            ],
            "schedule_slot": slot_label,
            "timestamp": datetime.now(IST).isoformat(),
        }
        await notification_manager.broadcast(notification)

    logger.info(f"🔔 Sent {len(stocks_with_recent_news)} notification(s) to frontend")
    logger.info(f"✅ Strategy Job COMPLETE — Slot: {slot_label}")
    logger.info(f"{'='*60}")


# =============================================================================
# Scheduler Startup
# =============================================================================


def start_strategy_scheduler() -> AsyncIOScheduler:
    """
    Parse schedule times from .env and register cron jobs for each slot.
    Returns the running AsyncIOScheduler instance.
    """
    logger.info("🔧 Initializing Strategy Scheduler...")

    times = [t.strip() for t in SCHEDULE_TIMES_STR.split(",") if t.strip()]
    scheduler = AsyncIOScheduler(timezone=ZoneInfo("Asia/Kolkata"))

    for idx, time_str in enumerate(times):
        try:
            parts = time_str.split(":")
            hour = int(parts[0])
            minute = int(parts[1])

            slot_label = f"slot_{idx + 1}_{hour:02d}{minute:02d}"

            trigger = CronTrigger(
                day_of_week="mon-fri",
                hour=hour,
                minute=minute,
                timezone=IST
            )

            scheduler.add_job(
                run_strategy_job,
                trigger=trigger,
                args=[idx, slot_label],
                id=f"strategy_job_{slot_label}",
                replace_existing=True,
            )

            vol_threshold = get_volume_threshold_for_slot(idx)
            logger.info(
                f"  ⏰ {hour:02d}:{minute:02d} Mon-Fri → {slot_label} "
                f"(min vol: {vol_threshold:,.0f})"
            )

        except Exception as e:
            logger.error(f"❌ Failed to parse schedule time '{time_str}': {e}")

    scheduler.start()
    logger.info("✅ Strategy Scheduler started successfully.")
    return scheduler


# =============================================================================
# Standalone runner (for testing)
# =============================================================================

if __name__ == "__main__":
    if os.getenv("ENVIRONMENT_OS", sys.platform) == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

    scheduler = start_strategy_scheduler()
    try:
        asyncio.get_event_loop().run_forever()
    except KeyboardInterrupt:
        logger.info("Program terminated by user.")
        scheduler.shutdown()
